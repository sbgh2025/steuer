import tkinter as tk
from tkinter import messagebox
import sqlite3
import openpyxl
from openpyxl.styles import Font

# Verbindung zur SQLite-Datenbank
conn = sqlite3.connect('steuern.db')
c = conn.cursor()

# Funktion zur Steuerübersicht
def steuer_uebersicht():
    # Einnahmen aus Selbstständigkeit
    c.execute("SELECT SUM(betrag) FROM tbl_ek_selbstst")
    einkommen_selbst = c.fetchone()[0] or 0.0

    # Ausgaben
    c.execute("SELECT SUM(betrag) FROM tbl_ag_vers")
    ausgaben_vers = c.fetchone()[0] or 0.0

    c.execute("SELECT SUM(betrag) FROM tbl_ag_buero")
    ausgaben_buero = c.fetchone()[0] or 0.0

    c.execute("SELECT SUM(betrag) FROM tbl_ag_porto")
    ausgaben_porto = c.fetchone()[0] or 0.0

    gesamt_ausgaben = ausgaben_vers + ausgaben_buero + ausgaben_porto
    gewinn_selbst = einkommen_selbst + gesamt_ausgaben

    # Kapitalerträge
    c.execute("SELECT SUM(betrag) FROM tbl_ek_ke")
    kapital_einnahmen = c.fetchone()[0] or 0.0

    # Abgeführte Kapitalsteuer aus Tabelle
    c.execute("SELECT SUM(betrag) FROM tbl_ag_abgeltst")
    abgeführte_steuer_db = -(c.fetchone()[0] or 0.0)  # negativ gespeichert → positiv umdrehen

    # Freibetrag
    try:
        freibetrag = float(sparer_pausch_var.get() or 801)
    except ValueError:
        messagebox.showerror("Fehler", "Ungültiger Sparerpauschbetrag.")
        return

    zu_versteuern = max(0, kapital_einnahmen - freibetrag)
    kapitalsteuer_soll = round(zu_versteuern * 0.25, 2)
    zahlbetrag = kapitalsteuer_soll - abgeführte_steuer_db

    steuer_rueckforderung = 0
    if kapital_einnahmen < freibetrag and abgeführte_steuer_db > 0:
        steuer_rueckforderung = abgeführte_steuer_db

    # Anzeige
    text = f"🧾 Selbständigkeit:\n"
    text += f"  Einnahmen: {einkommen_selbst:.2f} €\n"
    text += f"  Ausgaben:\n"
    text += f"    • Versicherung: {ausgaben_vers:.2f} €\n"
    text += f"    • Büro: {ausgaben_buero:.2f} €\n"
    text += f"    • Porto: {ausgaben_porto:.2f} €\n"
    text += f"  Gewinn: {gewinn_selbst:.2f} €\n\n"

    text += f"💰 Kapitalerträge:\n"
    text += f"  Brutto laut DB: {kapital_einnahmen:.2f} €\n"
    text += f"  Sparerpauschbetrag: {freibetrag:.2f} €\n"
    text += f"  Zu versteuern: {zu_versteuern:.2f} €\n"
    text += f"  Steuer (25%): {kapitalsteuer_soll:.2f} €\n"
    text += f"  Abgeführt: {abgeführte_steuer_db:.2f} €\n"

    if zahlbetrag > 0:
        text += f"🔺 Nachzahlung: {zahlbetrag:.2f} €"
    elif zahlbetrag < 0:
        text += f"🟢 Erstattung: {abs(zahlbetrag):.2f} €"
    else:
        text += f"✔️ Keine Nachzahlung oder Erstattung"

    if steuer_rueckforderung > 0:
        text += f"\n🔄 Rückforderung der Steuer: {steuer_rueckforderung:.2f} €"

    messagebox.showinfo("Steuerübersicht", text)

# Excel-Export
def export_to_excel():
    wb = openpyxl.Workbook()
    ws_eink = wb.active
    ws_eink.title = "Einkünfte Selbständig"
    ws_eink.append(["ID", "Datum", "Betrag (€)", "Bemerkung"])
    for col in "ABCD":
        ws_eink[f"{col}1"].font = Font(bold=True)
    c.execute("SELECT * FROM tbl_ek_selbstst")
    for row in c.fetchall():
        ws_eink.append(row)

    def add_sheet(title, table_name):
        ws = wb.create_sheet(title)
        ws.append(["ID", "Datum", "Betrag (€)", "Bemerkung"])
        for col in "ABCD":
            ws[f"{col}1"].font = Font(bold=True)
        c.execute(f"SELECT * FROM {table_name}")
        for row in c.fetchall():
            ws.append(row)

    # Weitere Tabellenblätter hinzufügen
    add_sheet("Ausgaben Versicherung", "tbl_ag_vers")
    add_sheet("Ausgaben Büro", "tbl_ag_buero")
    add_sheet("Ausgaben Porto", "tbl_ag_porto")
    add_sheet("Kapitalerträge", "tbl_ek_ke")
    add_sheet("Kapitalsteuer", "tbl_ag_abgeltst")

    # Übersicht berechnen
    c.execute("SELECT SUM(betrag) FROM tbl_ek_selbstst")
    einkommen_selbst = c.fetchone()[0] or 0.0

    c.execute("SELECT SUM(betrag) FROM tbl_ag_vers")
    ausgaben_vers = c.fetchone()[0] or 0.0

    c.execute("SELECT SUM(betrag) FROM tbl_ag_buero")
    ausgaben_buero = c.fetchone()[0] or 0.0

    c.execute("SELECT SUM(betrag) FROM tbl_ag_porto")
    ausgaben_porto = c.fetchone()[0] or 0.0

    gewinn_selbst = einkommen_selbst + ausgaben_vers + ausgaben_buero + ausgaben_porto

    c.execute("SELECT SUM(betrag) FROM tbl_ek_ke")
    kapital_einnahmen = c.fetchone()[0] or 0.0

    c.execute("SELECT SUM(betrag) FROM tbl_ag_abgeltst")
    abgefuehrt = -(c.fetchone()[0] or 0.0)

    freibetrag = float(sparer_pausch_var.get() or 801)
    zu_versteuern = max(0, kapital_einnahmen - freibetrag)
    kapitalsteuer_soll = round(zu_versteuern * 0.25, 2)
    zahlbetrag = kapitalsteuer_soll - abgefuehrt

    # Übersicht als eigenes Tabellenblatt
    ws_summary = wb.create_sheet("Übersicht")
    ws_summary.append(["Bereich", "Wert (€)"])
    ws_summary["A1"].font = ws_summary["B1"].font = Font(bold=True)

    ws_summary.append(["Einnahmen Selbständigkeit", einkommen_selbst])
    ws_summary.append(["Ausgaben Versicherung", ausgaben_vers])
    ws_summary.append(["Ausgaben Büro", ausgaben_buero])
    ws_summary.append(["Ausgaben Porto", ausgaben_porto])
    ws_summary.append(["Gewinn Selbständig", gewinn_selbst])
    ws_summary.append(["Kapitalerträge (brutto)", kapital_einnahmen])
    ws_summary.append(["Sparerpauschbetrag", freibetrag])
    ws_summary.append(["Zu versteuern", zu_versteuern])
    ws_summary.append(["Kapitalsteuer (25%)", kapitalsteuer_soll])
    ws_summary.append(["Abgeführt", abgefuehrt])
    ws_summary.append(["Nachzahlung (+) / Erstattung (-)", zahlbetrag])

    wb.save("steuerdaten_2025.xlsx")
    messagebox.showinfo("Excel Export", "Datei 'steuerdaten_2025.xlsx' wurde gespeichert.")

# GUI
root = tk.Tk()
root.title("Steuerverwaltung")
root.geometry("1200x200")  # Breite x Höhe in Pixeln


tk.Label(root, text="Sparerpauschbetrag (€):").grid(row=0, column=0)
sparer_pausch_var = tk.StringVar(value="801")
tk.Entry(root, textvariable=sparer_pausch_var).grid(row=0, column=1)

tk.Button(root, text="Übersicht anzeigen", command=steuer_uebersicht).grid(row=1, column=0, pady=5)
tk.Button(root, text="In Excel exportieren", command=export_to_excel).grid(row=2, column=0, pady=5)

root.mainloop()

conn.close()
